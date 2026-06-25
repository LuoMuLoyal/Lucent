from __future__ import annotations

import argparse
import json
from typing import Any, Iterator

from common import emit_error, emit_record, normalize_text, stable_uuid

try:
    from openpyxl import load_workbook
except ModuleNotFoundError:  # pragma: no cover - depends on local tooling
    load_workbook = None


# Headers exported by DrugDataBase/build_cn_master.py for the
# InstructionsClean sheet.
INSTRUCTION_FIELDS = [
    "instruction_id",
    "source_file",
    "source_row",
    "title",
    "title_url",
    "number_raw",
    "summary",
    "generic_name",
    "brand_name",
    "pinyin",
    "approval_raw",
    "approval_codes",
    "approval_conflict",
    "drug_category",
    "manufacturer",
    "drug_nature",
    "related_diseases",
    "properties",
    "ingredients",
    "indications",
    "package_spec",
    "adverse_reactions",
    "dosage",
    "contraindications",
    "precautions",
    "pregnancy_lactation",
    "pediatric_use",
    "geriatric_use",
    "drug_interactions",
    "pharmacology_toxicology",
    "pharmacokinetics",
    "storage",
    "validity_period",
    "merge_notes",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source-path", required=True)
    parser.add_argument("--limit", type=int)
    return parser.parse_args()


def parse_approval_codes(value: Any) -> list[str]:
    text = normalize_text(value)
    if text is None:
        return []
    return [code for code in text.split("|") if code]


def parse_source_row(value: Any) -> int | None:
    if value is None:
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def to_record(row_number: int, row: dict[str, Any]) -> dict[str, Any] | None:
    instruction_id = normalize_text(row.get("instruction_id"))
    if instruction_id is None:
        emit_error("Missing required instruction_id", row_number)
        return None

    approval_codes = parse_approval_codes(row.get("approval_codes"))

    record = {
        "id": stable_uuid("cn_medicine_leaflet", instruction_id),
        "instruction_id": instruction_id,
        "source_file": normalize_text(row.get("source_file")),
        "source_row": parse_source_row(row.get("source_row")),
        "title": normalize_text(row.get("title")),
        "title_url": normalize_text(row.get("title_url")),
        "number_raw": normalize_text(row.get("number_raw")),
        "summary": normalize_text(row.get("summary")),
        "generic_name": normalize_text(row.get("generic_name")),
        "brand_name": normalize_text(row.get("brand_name")),
        "pinyin": normalize_text(row.get("pinyin")),
        "approval_raw": normalize_text(row.get("approval_raw")),
        "approval_codes": approval_codes,
        "approval_conflict": normalize_text(row.get("approval_conflict")),
        "drug_category": normalize_text(row.get("drug_category")),
        "manufacturer": normalize_text(row.get("manufacturer")),
        "drug_nature": normalize_text(row.get("drug_nature")),
        "related_diseases": normalize_text(row.get("related_diseases")),
        "properties": normalize_text(row.get("properties")),
        "ingredients": normalize_text(row.get("ingredients")),
        "indications": normalize_text(row.get("indications")),
        "package_spec": normalize_text(row.get("package_spec")),
        "adverse_reactions": normalize_text(row.get("adverse_reactions")),
        "dosage": normalize_text(row.get("dosage")),
        "contraindications": normalize_text(row.get("contraindications")),
        "precautions": normalize_text(row.get("precautions")),
        "pregnancy_lactation": normalize_text(row.get("pregnancy_lactation")),
        "pediatric_use": normalize_text(row.get("pediatric_use")),
        "geriatric_use": normalize_text(row.get("geriatric_use")),
        "drug_interactions": normalize_text(row.get("drug_interactions")),
        "pharmacology_toxicology": normalize_text(
            row.get("pharmacology_toxicology")
        ),
        "pharmacokinetics": normalize_text(row.get("pharmacokinetics")),
        "storage": normalize_text(row.get("storage")),
        "validity_period": normalize_text(row.get("validity_period")),
        "merge_notes": normalize_text(row.get("merge_notes")),
    }
    return record


def iter_xlsx_rows(source_path: str) -> Iterator[tuple[int, dict[str, Any]]]:
    if load_workbook is None:
        raise SystemExit(
            "openpyxl is required for .xlsx imports. "
            "Run `pip install -r scripts/import/medicine/requirements.txt` "
            "or export the sheet to CSV and use --source-path <file.csv>."
        )

    workbook = load_workbook(source_path, read_only=True)

    if "InstructionsClean" not in workbook.sheetnames:
        raise SystemExit(
            "Expected sheet 'InstructionsClean' was not found in the workbook"
        )

    worksheet = workbook["InstructionsClean"]
    worksheet_rows = worksheet.iter_rows(values_only=True)
    header = next(worksheet_rows, None)
    if header is None:
        raise SystemExit("Workbook is empty")

    header_names = [normalize_text(cell) or "" for cell in header]
    for row_number, values in enumerate(worksheet_rows, start=2):
        yield row_number, dict(zip(header_names, values))


def main() -> None:
    args = parse_args()
    rows = iter_xlsx_rows(args.source_path)

    emitted = 0
    for row_number, row in rows:
        record = to_record(row_number, row)
        if record is None:
            continue

        emit_record(record)
        emitted += 1
        if args.limit is not None and emitted >= args.limit:
            break


if __name__ == "__main__":
    main()
