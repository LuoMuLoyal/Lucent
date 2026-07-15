from __future__ import annotations

import argparse
import csv
from pathlib import Path
from typing import Any, Iterator

from common import build_search_text, emit_error, emit_record, normalize_text, stable_uuid

try:
    from openpyxl import load_workbook
except ModuleNotFoundError:  # pragma: no cover - depends on local tooling
    load_workbook = None


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source-path", required=True)
    parser.add_argument("--limit", type=int)
    return parser.parse_args()


def parse_pipe_list(value: Any) -> list[str]:
    text = normalize_text(value)
    if text is None:
        return []
    return [item for item in text.split("|") if item]


def parse_int_or_none(value: Any) -> int | None:
    if value is None:
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def to_record(row_number: int, row: dict[str, Any]) -> dict[str, Any] | None:
    name = normalize_text(row.get("product_name"))
    if name is None:
        emit_error("Missing required product_name", row_number)
        return None

    package_spec = normalize_text(row.get("package_spec"))
    approval_number = normalize_text(row.get("approval_number"))
    manufacturer = normalize_text(row.get("manufacturer"))
    national_drug_code = normalize_text(row.get("national_drug_code"))

    if approval_number is not None:
        product_id = stable_uuid(
            "cn_medicine_product",
            approval_number,
            package_spec,
            manufacturer,
        )
    else:
        product_id = stable_uuid(
            "cn_medicine_product",
            name,
            package_spec,
            manufacturer,
            national_drug_code,
            row_number,
        )

    record = {
        "id": product_id,
        "source_name": "chinese_drug_data_master",
        "source_row_number": row_number,
        "name": name,
        "image_url": normalize_text(row.get("image_url")),
        "price_text": normalize_text(row.get("price")),
        "package_spec": package_spec,
        "approval_number": approval_number,
        "manufacturer": manufacturer,
        "drug_type": normalize_text(row.get("drug_type")),
        "main_category": normalize_text(row.get("main_category")),
        "subcategory": normalize_text(row.get("subcategory")),
        "source_url": normalize_text(row.get("detail_url")),
        "brand_name": normalize_text(row.get("brand_name")),
        "ingredients": normalize_text(row.get("ingredients")),
        "properties": normalize_text(row.get("properties")),
        "indications": normalize_text(row.get("indications")),
        "dosage": normalize_text(row.get("dosage")),
        "adverse_reactions": normalize_text(row.get("adverse_reactions")),
        "contraindications": normalize_text(row.get("contraindications")),
        "precautions": normalize_text(row.get("precautions")),
        "pediatric_use": normalize_text(row.get("pediatric_use")),
        "geriatric_use": normalize_text(row.get("geriatric_use")),
        "pregnancy_lactation": normalize_text(row.get("pregnancy_lactation")),
        "pharmacology_toxicology": normalize_text(
            row.get("pharmacology_toxicology")
        ),
        "drug_interactions": normalize_text(row.get("drug_interactions")),
        "pharmacokinetics": normalize_text(row.get("pharmacokinetics")),
        "overdose": normalize_text(row.get("overdose")),
        "storage": normalize_text(row.get("storage")),
        "validity_period": normalize_text(row.get("validity_period")),
        "barcode": normalize_text(row.get("barcode")),
        "national_drug_code": national_drug_code,
        "search_text": build_search_text(
            [
                name,
                normalize_text(row.get("brand_name")),
                manufacturer,
                normalize_text(row.get("manufacturer_normalized")),
                approval_number,
                normalize_text(row.get("barcode")),
                national_drug_code,
            ]
        ),
        "image_url_cleaned": normalize_text(row.get("image_url_cleaned")),
        "manufacturer_normalized": normalize_text(row.get("manufacturer_normalized")),
        "approval_codes": parse_pipe_list(row.get("approval_codes")),
        "best_match_type": normalize_text(row.get("best_match_type")),
        "best_match_score": parse_int_or_none(row.get("best_match_score")),
        "top_candidate_ids": parse_pipe_list(row.get("top_candidate_ids")),
        "top_candidate_scores": [
            int(score)
            for score in parse_pipe_list(row.get("top_candidate_scores"))
            if score.isdigit()
        ],
        "candidate_count": parse_int_or_none(row.get("candidate_count")),
        "match_quality_overall": parse_int_or_none(row.get("match_quality_overall")),
        "match_quality_approval": parse_int_or_none(row.get("match_quality_approval")),
        "match_quality_name": parse_int_or_none(row.get("match_quality_name")),
        "match_quality_maker": parse_int_or_none(row.get("match_quality_maker")),
        "match_quality_leaflet": parse_int_or_none(row.get("match_quality_leaflet")),
        "match_quality_penalty": parse_int_or_none(row.get("match_quality_penalty")),
        "match_quality_notes": parse_pipe_list(row.get("match_quality_notes")),
        "extras": None,
    }
    return record


def iter_csv_rows(source_path: str) -> Iterator[tuple[int, dict[str, Any]]]:
    with open(source_path, encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        for row_number, row in enumerate(reader, start=2):
            normalized_row = {
                (normalize_text(key) or ""): value for key, value in row.items()
            }
            yield row_number, normalized_row


def iter_xlsx_rows(source_path: str) -> Iterator[tuple[int, dict[str, Any]]]:
    if load_workbook is None:
        raise SystemExit(
            "openpyxl is required for .xlsx imports. "
            "Run `pip install -r scripts/import/medicine/requirements.txt` "
            "or export the sheet to CSV and use --source-path <file.csv>."
        )

    workbook = load_workbook(source_path, read_only=True)

    # Prefer the enriched master sheet; fall back to the legacy raw sheet.
    if "ProductsEnriched" in workbook.sheetnames:
        sheet_name = "ProductsEnriched"
    elif "总的" in workbook.sheetnames:
        sheet_name = "总的"
    else:
        raise SystemExit(
            "Expected sheet 'ProductsEnriched' or '总的' was not found in the workbook"
        )

    worksheet = workbook[sheet_name]
    worksheet_rows = worksheet.iter_rows(values_only=True)
    header = next(worksheet_rows, None)
    if header is None:
        raise SystemExit("Workbook is empty")

    header_names = [normalize_text(cell) or "" for cell in header]
    for row_number, values in enumerate(worksheet_rows, start=2):
        yield row_number, dict(zip(header_names, values))


def main() -> None:
    args = parse_args()
    source_suffix = Path(args.source_path).suffix.lower()
    if source_suffix == ".csv":
        rows = iter_csv_rows(args.source_path)
    else:
        rows = iter_xlsx_rows(args.source_path)

    emitted = 0
    for row_number, row in rows:
        record = to_record(row_number, row)
        if record is None:
            continue

        emit_record(record)
        emitted += 1
        if args.limit is not None and emitted >= args.limit:
            break


if __name__ == "__main__":
    main()
