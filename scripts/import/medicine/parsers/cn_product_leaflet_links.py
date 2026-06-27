from __future__ import annotations

import argparse
from typing import Any, Iterator

from common import emit_error, emit_record, normalize_text, stable_uuid

try:
    from openpyxl import load_workbook
except ModuleNotFoundError:  # pragma: no cover - depends on local tooling
    load_workbook = None


# Headers exported by DrugDataBase/ChineseDrugData_Master_V2/build_master_v2.py
# for the ProductInstructionLinks sheet.
LINK_FIELDS = [
    "product_id",
    "instruction_id",
    "match_type",
    "match_key",
    "match_score",
    "product_name",
    "instruction_generic_name",
    "product_manufacturer",
    "instruction_manufacturer",
    "product_package_spec",
    "instruction_package_spec",
]

# Headers from ProductsEnriched needed to compute the same stable product id
# used by cn_products.py and to identify the best matched instruction.
PRODUCT_LOOKUP_FIELDS = [
    "product_id",
    "product_name",
    "package_spec",
    "approval_number",
    "manufacturer",
    "national_drug_code",
    "best_instruction_id",
    "best_match_type",
    "best_match_score",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source-path", required=True)
    parser.add_argument("--limit", type=int)
    return parser.parse_args()


def compute_product_id(
    product_name: str | None,
    package_spec: str | None,
    approval_number: str | None,
    manufacturer: str | None,
    national_drug_code: str | None,
    source_row_number: int,
) -> str:
    if approval_number is not None:
        return stable_uuid(
            "cn_medicine_product",
            approval_number,
            package_spec,
            manufacturer,
        )

    return stable_uuid(
        "cn_medicine_product",
        product_name,
        package_spec,
        manufacturer,
        national_drug_code,
        source_row_number,
    )


def parse_int_or_none(value: Any) -> int | None:
    if value is None:
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def read_products_enriched(source_path: str) -> dict[str, dict[str, Any]]:
    if load_workbook is None:
        raise SystemExit(
            "openpyxl is required for .xlsx imports. "
            "Run `pip install -r scripts/import/medicine/requirements.txt`"
        )

    workbook = load_workbook(source_path, read_only=True)
    if "ProductsEnriched" not in workbook.sheetnames:
        raise SystemExit(
            "Expected sheet 'ProductsEnriched' was not found in the workbook"
        )

    worksheet = workbook["ProductsEnriched"]
    worksheet_rows = worksheet.iter_rows(values_only=True)
    header = next(worksheet_rows, None)
    if header is None:
        raise SystemExit("ProductsEnriched sheet is empty")

    header_names = [normalize_text(cell) or "" for cell in header]
    products: dict[str, dict[str, Any]] = {}
    for row_number, values in enumerate(worksheet_rows, start=2):
        row = dict(zip(header_names, values))
        product_id = normalize_text(row.get("product_id"))
        if product_id is None:
            continue
        products[product_id] = {
            "source_row_number": row_number,
            "product_name": normalize_text(row.get("product_name")),
            "package_spec": normalize_text(row.get("package_spec")),
            "approval_number": normalize_text(row.get("approval_number")),
            "manufacturer": normalize_text(row.get("manufacturer")),
            "national_drug_code": normalize_text(row.get("national_drug_code")),
            "best_instruction_id": normalize_text(row.get("best_instruction_id")),
            "best_match_type": normalize_text(row.get("best_match_type")),
            "best_match_score": parse_int_or_none(row.get("best_match_score")),
        }

    return products


def build_link_id(
    product_id: str,
    leaflet_id: str,
    approval_code: str | None,
) -> str:
    return stable_uuid(
        "cn_product_leaflet_link",
        product_id,
        leaflet_id,
        approval_code,
    )


def build_best_match_record(
    product: dict[str, Any],
) -> dict[str, Any] | None:
    best_instruction_id = product.get("best_instruction_id")
    if best_instruction_id is None:
        return None

    product_id = compute_product_id(
        product["product_name"],
        product["package_spec"],
        product["approval_number"],
        product["manufacturer"],
        product["national_drug_code"],
        product["source_row_number"],
    )
    leaflet_id = stable_uuid("cn_medicine_leaflet", best_instruction_id)

    return {
        "id": build_link_id(product_id, leaflet_id, None),
        "product_id": product_id,
        "leaflet_id": leaflet_id,
        "approval_code": None,
        "match_type": product.get("best_match_type"),
        "match_score": product.get("best_match_score"),
        "is_best_match": True,
    }


def merge_link_records(
    existing: dict[str, Any],
    incoming: dict[str, Any],
) -> dict[str, Any]:
    """Keep the highest-scoring record while preserving is_best_match."""
    if incoming["is_best_match"]:
        return incoming
    if existing["is_best_match"]:
        existing_score = existing.get("match_score") or 0
        incoming_score = incoming.get("match_score") or 0
        if incoming_score > existing_score:
            # Prefer higher score but keep the best-match flag.
            return {**incoming, "is_best_match": True}
        return existing

    existing_score = existing.get("match_score") or 0
    incoming_score = incoming.get("match_score") or 0
    return incoming if incoming_score > existing_score else existing


def to_record(
    row_number: int,
    row: dict[str, Any],
    products: dict[str, dict[str, Any]],
) -> dict[str, Any] | None:
    source_product_id = normalize_text(row.get("product_id"))
    instruction_id = normalize_text(row.get("instruction_id"))
    if source_product_id is None:
        emit_error("Missing required product_id", row_number)
        return None
    if instruction_id is None:
        emit_error("Missing required instruction_id", row_number)
        return None

    product = products.get(source_product_id)
    if product is None:
        emit_error(
            f"Product {source_product_id} not found in ProductsEnriched",
            row_number,
        )
        return None

    product_id = compute_product_id(
        product["product_name"],
        product["package_spec"],
        product["approval_number"],
        product["manufacturer"],
        product["national_drug_code"],
        product["source_row_number"],
    )
    leaflet_id = stable_uuid("cn_medicine_leaflet", instruction_id)

    match_score = parse_int_or_none(row.get("match_score"))
    match_type = normalize_text(row.get("match_type"))
    match_key = normalize_text(row.get("match_key"))
    approval_code = match_key if match_type == "exact_code" else None

    is_best_match = product["best_instruction_id"] == instruction_id

    return {
        "id": build_link_id(product_id, leaflet_id, approval_code),
        "product_id": product_id,
        "leaflet_id": leaflet_id,
        "approval_code": approval_code,
        "match_type": match_type,
        "match_score": match_score,
        "is_best_match": is_best_match,
    }


def iter_xlsx_rows(
    source_path: str,
) -> Iterator[tuple[int, dict[str, Any]]]:
    if load_workbook is None:
        raise SystemExit(
            "openpyxl is required for .xlsx imports. "
            "Run `pip install -r scripts/import/medicine/requirements.txt` "
            "or export the sheet to CSV and use --source-path <file.csv>."
        )

    workbook = load_workbook(source_path, read_only=True)
    if "ProductInstructionLinks" not in workbook.sheetnames:
        raise SystemExit(
            "Expected sheet 'ProductInstructionLinks' was not found in the workbook"
        )

    worksheet = workbook["ProductInstructionLinks"]
    worksheet_rows = worksheet.iter_rows(values_only=True)
    header = next(worksheet_rows, None)
    if header is None:
        raise SystemExit("ProductInstructionLinks sheet is empty")

    header_names = [normalize_text(cell) or "" for cell in header]
    for row_number, values in enumerate(worksheet_rows, start=2):
        yield row_number, dict(zip(header_names, values))


def main() -> None:
    args = parse_args()
    products = read_products_enriched(args.source_path)
    rows = iter_xlsx_rows(args.source_path)

    # First pass: ensure every product has a link to its best instruction,
    # regardless of whether it came from exact_code or fuzzy_name matching.
    links: dict[tuple[str, str], dict[str, Any]] = {}
    for product in products.values():
        best_record = build_best_match_record(product)
        if best_record is None:
            continue
        key = (best_record["product_id"], best_record["leaflet_id"])
        links[key] = best_record

    # Second pass: add exact-code candidate links from ProductInstructionLinks.
    # If a candidate is the product's best instruction, the flag is preserved.
    emitted = 0
    for row_number, row in rows:
        record = to_record(row_number, row, products)
        if record is None:
            continue

        key = (record["product_id"], record["leaflet_id"])
        existing = links.get(key)
        if existing is None:
            links[key] = record
        else:
            links[key] = merge_link_records(existing, record)

        emitted += 1
        if args.limit is not None and emitted >= args.limit:
            break

    emitted = 0
    for record in links.values():
        if args.limit is not None and emitted >= args.limit:
            break
        emit_record(record)
        emitted += 1


if __name__ == "__main__":
    main()
