from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import load_workbook

from common import build_search_text, emit_error, emit_record, normalize_text


SHEET_NAME = "中国食物分类表"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source-path", required=True)
    parser.add_argument("--limit", type=int)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    workbook = load_workbook(Path(args.source_path), read_only=True, data_only=True)
    if SHEET_NAME not in workbook.sheetnames:
        raise SystemExit(f"Missing sheet: {SHEET_NAME}")

    sheet = workbook[SHEET_NAME]
    rows = sheet.iter_rows(values_only=True)
    next(rows, None)
    header = [
        normalize_text(cell).replace("  ", " ")
        if normalize_text(cell) is not None
        else None
        for cell in next(rows, [])
    ]
    index = {name: position for position, name in enumerate(header) if name}

    emitted = 0
    for row_number, row in enumerate(rows, start=3):
        code = normalize_text(row[index["编码"]]) if "编码" in index else None
        name = normalize_text(row[index["名称"]]) if "名称" in index else None
        if code is None or name is None:
          emit_error("Missing category code or name", row_number)
          continue

        parent_code = (
            normalize_text(row[index["父级编码"]]) if "父级编码" in index else None
        )
        emit_record(
            {
                "code": code,
                "source_row_number": row_number,
                "parent_code": parent_code,
                "name": name,
                "level": 1 if len(code) == 2 else 2 if len(code) == 3 else 3,
                "search_text": build_search_text([code, name, parent_code]),
            }
        )
        emitted += 1

        if args.limit is not None and emitted >= args.limit:
            break


if __name__ == "__main__":
    main()
