from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import load_workbook

from common import (
    build_search_text,
    emit_error,
    emit_record,
    normalize_name,
    normalize_text,
    stable_uuid,
    to_float,
)


SHEET_NAME = "中国食物成分表"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source-path", required=True)
    parser.add_argument("--limit", type=int)
    return parser.parse_args()


def value(row: tuple[object, ...], index: dict[str, int], column: str):
    position = index.get(column)
    if position is None or position >= len(row):
        return None
    return row[position]


def main() -> None:
    args = parse_args()
    workbook = load_workbook(Path(args.source_path), read_only=True, data_only=True)
    if SHEET_NAME not in workbook.sheetnames:
        raise SystemExit(f"Missing sheet: {SHEET_NAME}")

    sheet = workbook[SHEET_NAME]
    rows = sheet.iter_rows(values_only=True)
    next(rows, None)
    header = [
        normalize_text(cell).replace("  ", " ")
        if normalize_text(cell) is not None
        else None
        for cell in next(rows, [])
    ]
    index = {name: position for position, name in enumerate(header) if name}

    emitted = 0
    for row_number, row in enumerate(rows, start=3):
        serial_number = value(row, index, "序号")
        name = normalize_text(value(row, index, "名 称"))
        normalized_name = normalize_name(name)
        if name is None or normalized_name is None:
            emit_error("Missing food item name", row_number)
            continue

        primary_category_code = normalize_text(value(row, index, "一级分类编码"))
        secondary_category_code = normalize_text(value(row, index, "二级分类编码"))

        emit_record(
            {
                "id": stable_uuid("food_item", serial_number, name),
                "source_row_number": row_number,
                "source_serial_number": int(serial_number)
                if isinstance(serial_number, (int, float))
                else None,
                "name": name,
                "normalized_name": normalized_name,
                "search_text": build_search_text(
                    [name, normalized_name, primary_category_code, secondary_category_code]
                ),
                "aliases": [normalized_name] if normalized_name != name else [],
                "primary_category_code": primary_category_code,
                "secondary_category_code": secondary_category_code,
                "edible_portion_percent": to_float(value(row, index, "食部(%)")),
                "water_g": to_float(value(row, index, "水分(g)")),
                "energy_kcal": to_float(value(row, index, "能量（千卡）")),
                "energy_kj": to_float(value(row, index, "能量（千焦）")),
                "protein_g": to_float(value(row, index, "蛋白质(g)")),
                "fat_g": to_float(value(row, index, "脂肪(g)")),
                "carbohydrate_g": to_float(value(row, index, "碳水化物(g)")),
                "fiber_g": to_float(value(row, index, "膳食纤维(g)")),
                "cholesterol_mg": to_float(value(row, index, "胆固醇(mg)")),
                "calcium_mg": to_float(value(row, index, "钙(mg)")),
                "phosphorus_mg": to_float(value(row, index, "磷(mg)")),
                "potassium_mg": to_float(value(row, index, "钾(mg)")),
                "sodium_mg": to_float(value(row, index, "钠(mg)")),
                "magnesium_mg": to_float(value(row, index, "镁(mg)")),
                "iron_mg": to_float(value(row, index, "铁(mg)")),
                "zinc_mg": to_float(value(row, index, "锌(mg)")),
                "selenium_mg": to_float(value(row, index, "硒(mg)")),
                "copper_mg": to_float(value(row, index, "铜(mg)")),
                "manganese_mg": to_float(value(row, index, "锰(mg)")),
                "iodine_mg": to_float(value(row, index, "碘(mg)")),
                "vitamin_a_mcg_re": to_float(value(row, index, "维生素A(μgRE)")),
                "thiamin_mg": to_float(value(row, index, "硫胺素(mg)")),
                "riboflavin_mg": to_float(value(row, index, "核黄素(mg)")),
                "vitamin_b6_mg": to_float(value(row, index, "维生素B6(mg)")),
                "vitamin_b12_mg": to_float(value(row, index, "维生素B12(mg)")),
                "folate_ug": to_float(value(row, index, "叶酸(ug)")),
                "niacin_mg": to_float(value(row, index, "尼克酸/烟酸(mg)")),
                "vitamin_c_mg": to_float(value(row, index, "维生素C(mg)")),
                "vitamin_e_mg": to_float(value(row, index, "维生素E(mg)")),
                "carotene_mcg": to_float(value(row, index, "胡萝卜素(μg)")),
                "retinol_mcg": to_float(value(row, index, "视黄醇(μg)")),
                "alpha_vitamin_e_mg": to_float(value(row, index, "α-维生素E(mg)")),
                "extras": {
                    "ash_g": to_float(value(row, index, "灰分(g)")),
                },
            }
        )
        emitted += 1

        if args.limit is not None and emitted >= args.limit:
            break


if __name__ == "__main__":
    main()
